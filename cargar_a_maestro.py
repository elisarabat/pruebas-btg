# -*- coding: utf-8 -*-
"""
Script para cargar datos desde archivos Excel fuente hacia un archivo maestro.
- Evita duplicados por Rut + Fecha de emisión.
- Mapea columnas aunque los nombres no coincidan exactamente.
"""

import pandas as pd
import sys
import os
import unicodedata
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


# Columnas del maestro que se calculan desde otras (no vienen del archivo fuente)
COLUMNAS_CALCULADAS = {
    "ID Blotter": "N° OP",  # dígitos de N° OP antes de la primera letra
    "Dif. Tasa": "Tasa Arriendo o Compra - Tasa Venta",
}
# Columnas que se rellenan desde columnas del fuente (fechas y columna anterior a 2ª fecha)
COLUMNAS_DESDE_FECHA = ("VPN 1ra fecha", "VPN 2da fecha", "Precio -1UF")
# Hojas del Excel fuente: datos principales y hoja auxiliar
HOJA_VALO = "Valo"
HOJA_BASE = "Base"
# Fila donde están los nombres de columna en el archivo fuente (0 = primera fila)
FILA_ENCABEZADO_VALO = 2   # fila 3 en Excel (o ver USAR_FILAS_2_Y_3_VALO)
FILA_ENCABEZADO_BASE = 0   # fila 1 en Excel
# Si True, en la hoja Valo los nombres de columna se construyen desde filas 2 y 3:
# para cada columna se usa el valor de la fila 3 si no está vacío, si no el de la fila 2.
# Así las columnas que solo tienen título en la fila 2 dejan de salir como "Unnamed".
USAR_FILAS_2_Y_3_VALO = True
# Llave para enlazar filas de Valo con Base (debe existir en ambas). Si en Base tiene otro nombre, ponlo en COLUMNA_LLAVE_EN_BASE.
LLAVE_PARA_BASE = "Rut"
COLUMNA_LLAVE_EN_BASE = None  # None = buscar columna con mismo nombre normalizado en Base
# 3 columnas del maestro que se rellenan desde la hoja Base: { nombre_en_maestro: nombre_columna_en_Base }
COLUMNAS_DESDE_BASE = {
    "Fecha de emisión": "Fecha de suscripción",
    "Tasa Arriendo o Compra": "Tasa anual de emisión",
    "Tasa Venta": "Tasa anual de endoso",
}

# Mapeo por índice cuando la columna en el fuente tiene nombre vacío (Unnamed).
# Si USAR_FILAS_2_Y_3_VALO = True, muchas de estas columnas ya tendrán nombre (desde la fila 2)
# y no hará falta listarlas aquí. Este dict es fallback por si alguna sigue sin coincidir.
# Clave = columna en el maestro; valor = índice 0-based en la hoja Valo.
COLUMNAS_POR_INDICE_FUENTE = {
    "Monto Crédito + Cap (UF)": 8,
    "Fecha 1er Aporte": 15,
    "Fecha Last Aporte": 16,
    "Fecha Corte": 17,
    "Tasacion": 23,
    "Precio Venta/Tasación": 24,
    "Monto dividendo": 25,
    "Div/Renta": 26,
    "Carga Financiera": 27,
}

# Columnas que son fechas: se convierten a datetime al leer y se formatean al escribir
COLUMNAS_FECHA = (
    "Fecha de compra",
    "Fecha de emisión",
    "Fecha 1er Aporte",
    "Fecha Last Aporte",
    "Fecha Corte",
)

# Columnas del archivo maestro (orden exacto)
COLUMNAS_MAESTRO = [
    "Fecha de compra",
    "N°",
    "N° OP",
    "ID Blotter",
    "Apellido Paterno",
    "Apellido materno",
    "Nombres",
    "Rut",
    "DV",
    "Fecha de emisión",
    "Monto Crédito + Cap (UF)",
    "Subsidio",
    "Pie",
    "Valor Vivienda",
    "Morosidad",
    "N° Cuotas",
    "Cuota Mes",
    "Tasa Arriendo o Compra",
    "Fecha 1er Aporte",
    "Fecha Last Aporte",
    "Fecha Corte",
    "VPN 1ra fecha",   # valor de la columna del fuente cuya cabecera es la primera fecha (orden cronológico)
    "VPN 2da fecha",   # valor de la columna cuya cabecera es la segunda fecha (siempre > primera)
    "Precio -1UF",     # valor de la columna de la 2ª fecha menos 1 (ej. col 09-02-2024 - 1)
    "Saldo insoluto Teorico al 31-07-2019",
    "Tasacion",
    "Precio Venta/Tasación",
    "Tasa Venta",
    "Dif. Tasa",
    "Monto dividendo",
    "Div/Renta",
    "Carga Financiera",
    "Dirección",
    "Comuna",
]


def extraer_id_blotter_desde_n_op(val) -> str:
    """
    Extrae el ID Blotter desde el valor de N° OP: todos los dígitos
    al inicio, antes de la primera letra. Ej: '12345ABC' -> '12345', '987' -> '987'.
    """
    if pd.isna(val):
        return ""
    s = str(val).strip()
    numeros = []
    for c in s:
        if c.isdigit():
            numeros.append(c)
        else:
            break
    return "".join(numeros)


def _parsear_nombre_como_fecha(nombre) -> pd.Timestamp | None:
    """Intenta parsear el nombre de columna como fecha. Devuelve Timestamp o None."""
    if pd.isna(nombre) or str(nombre).strip() == "":
        return None
    s = str(nombre).strip()
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%m-%d-%Y", "%d-%m-%y", "%d/%m/%y",
                "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M:%S.%f"):
        try:
            return pd.to_datetime(s, format=fmt)
        except (ValueError, TypeError):
            continue
    try:
        return pd.to_datetime(s, dayfirst=True)
    except (ValueError, TypeError):
        return None
    except (ValueError, TypeError):
        return None


def obtener_columnas_fecha_vpn(df_fuente: pd.DataFrame, columnas_excluir: set = None) -> tuple:
    """
    Encuentra columnas del fuente cuyo encabezado es una fecha.
    Devuelve (col_primera_fecha, col_segunda_fecha) ordenadas por fecha ascendente
    (la segunda fecha siempre será mayor que la primera). Si hay menos de 2, devuelve None.
    columnas_excluir: columnas ya mapeadas a otras columnas del maestro (no son VPN).
    """
    excluir = columnas_excluir or set()
    candidatos = []
    for col in df_fuente.columns:
        if col in excluir:
            continue
        d = _parsear_nombre_como_fecha(col)
        if d is not None and not pd.isna(d):
            candidatos.append((col, d))
    candidatos.sort(key=lambda x: x[1])
    if len(candidatos) >= 2:
        return candidatos[0][0], candidatos[1][0]
    if len(candidatos) == 1:
        return candidatos[0][0], None
    return None, None


def normalizar_nombre(col: str) -> str:
    """Normaliza nombre de columna para comparación: minúsculas, sin acentos, sin espacios extra."""
    if pd.isna(col) or not isinstance(col, str):
        return ""
    s = str(col).strip().lower()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    # Unificar variantes de ordinales (1.er, 1 er, 1ª -> 1er)
    s = s.replace("1.er", "1er").replace("1 er", "1er").replace("1. er", "1er")
    s = s.replace("1.ª", "1").replace("1ª", "1")
    # Unificar signos +/− (full-width, minus sign Unicode) y colapsar espacios
    s = s.replace("＋", "+").replace("－", "-").replace("−", "-")
    s = " ".join(s.split())
    return s


def _buscar_columna_en_df(df: pd.DataFrame, nombre: str) -> str | None:
    """Devuelve el nombre real de la columna en df que coincide con nombre (normalizado), o None."""
    if nombre is None or df is None or df.empty:
        return None
    norm = normalizar_nombre(nombre)
    for c in df.columns:
        if normalizar_nombre(c) == norm:
            return c
    return None


def _maestro_a_columnas_estandar(df_maestro: pd.DataFrame) -> pd.DataFrame:
    """
    Convierte el DataFrame leído del maestro a uno con columnas exactamente COLUMNAS_MAESTRO,
    mapeando por nombre normalizado. Así no se pierden datos de filas ya existentes cuando
    el Excel tiene nombres con espacios extra, acentos distintos, etc.
    """
    out = pd.DataFrame(index=df_maestro.index)
    for col_maestro in COLUMNAS_MAESTRO:
        col_real = _buscar_columna_en_df(df_maestro, col_maestro)
        if col_real is not None:
            out[col_maestro] = df_maestro[col_real].values
        else:
            out[col_maestro] = pd.NA
    return out


def _rut_valor_a_str(val) -> str:
    """Convierte un valor de RUT a string limpio, manejando floats (12345678.0 → '12345678')."""
    if pd.isna(val):
        return "nan"
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    s = str(val).strip()
    if s.replace(".", "", 1).isdigit() and "." in s and not s.startswith("0"):
        try:
            return str(int(float(s)))
        except (ValueError, OverflowError):
            pass
    return s.replace(".", "")


def _normalizar_fecha_str(val) -> str:
    """Normaliza un valor de fecha a string YYYY-MM-DD para comparación de duplicados."""
    if pd.isna(val):
        return ""
    try:
        dt = pd.to_datetime(val, dayfirst=True, errors="coerce")
        if pd.isna(dt):
            return ""
        return dt.strftime("%Y-%m-%d")
    except Exception:
        return ""


def _normalizar_rut_para_merge(val) -> str:
    """
    Normaliza un RUT para comparación: quita puntos y deja formato '12345678-9'.
    Acepta RUT con guion y DV (ej. '12.345.678-9') o solo número.
    """
    if pd.isna(val):
        return ""
    if isinstance(val, float) and val == int(val):
        return str(int(val)).upper()
    s = str(val).strip().replace(".", "").upper()
    return s


def _construir_llave_rut_desde_separados(rut_serie: pd.Series, dv_serie: pd.Series) -> pd.Series:
    """
    Construye una llave RUT normalizada desde columnas Rut y DV separadas (hoja Valo).
    Formato resultado: '12345678-9' (sin puntos, DV en mayúscula).
    """
    rut = rut_serie.apply(_rut_valor_a_str)
    dv = dv_serie.astype(str).str.strip().str.upper()
    llave = rut + "-" + dv
    llave = llave.replace(["nan-nan", "nan-NAN", "NaN-NaN", "nan-NaN"], "")
    return llave


def rellenar_desde_hoja_base(df_out: pd.DataFrame, df_base: pd.DataFrame) -> None:
    """
    Rellena las columnas definidas en COLUMNAS_DESDE_BASE con los valores de la hoja Base,
    enlazando por la columna LLAVE_PARA_BASE (ej. Rut). Si hay varias filas en Base con la misma
    llave, se usa la primera. Modifica df_out in-place.
    """
    if not COLUMNAS_DESDE_BASE or df_base is None or df_base.empty:
        return
    llave_maestro = LLAVE_PARA_BASE
    if llave_maestro not in df_out.columns:
        return
    # Nombre de la columna llave en Base
    llave_base = COLUMNA_LLAVE_EN_BASE if COLUMNA_LLAVE_EN_BASE else _buscar_columna_en_df(df_base, llave_maestro)
    if llave_base is None:
        return
    # Columnas de Base que necesitamos (maestro -> nombre en Base)
    cols_base = [llave_base]
    mapeo_base_a_maestro = {}
    for col_maestro, col_base_nombre in COLUMNAS_DESDE_BASE.items():
        col_base_real = _buscar_columna_en_df(df_base, col_base_nombre) if col_base_nombre else None
        if col_base_real and col_maestro in df_out.columns:
            cols_base.append(col_base_real)
            mapeo_base_a_maestro[col_base_real] = col_maestro
    cols_base = list(dict.fromkeys(cols_base))
    df_lookup = df_base[cols_base].drop_duplicates(subset=[llave_base], keep="first")

    # Normalizar llave para el merge: en Valo Rut y DV suelen estar separados; en Base vienen "12345678-9"
    if llave_maestro == "Rut" and "DV" in df_out.columns:
        df_out["_llave_merge_"] = _construir_llave_rut_desde_separados(df_out["Rut"], df_out["DV"])
    else:
        df_out["_llave_merge_"] = df_out[llave_maestro].apply(_normalizar_rut_para_merge)

    df_lookup["_llave_merge_"] = df_lookup[llave_base].apply(_normalizar_rut_para_merge)
    merged = df_out[["_llave_merge_"]].merge(
        df_lookup.drop(columns=[llave_base]),
        on="_llave_merge_",
        how="left",
        suffixes=("", "_base"),
    )
    for col_base_real, col_maestro in mapeo_base_a_maestro.items():
        if col_base_real in merged.columns:
            base_vals = pd.Series(merged[col_base_real].values, index=df_out.index)
            mask_valido = base_vals.notna()
            df_out.loc[mask_valido, col_maestro] = base_vals[mask_valido]
    df_out.drop(columns=["_llave_merge_"], inplace=True)


def mapear_columnas_fuente_a_maestro(df_fuente: pd.DataFrame) -> dict:
    """
    Devuelve un diccionario: nombre_columna_maestro -> nombre_columna_en_fuente.
    Si no hay coincidencia, la columna maestro no estará en el dict (quedará NaN).
    """
    nombres_fuente = list(df_fuente.columns)
    normalizados_fuente = {normalizar_nombre(c): c for c in nombres_fuente}

    mapeo = {}
    for col_maestro in COLUMNAS_MAESTRO:
        norm_maestro = normalizar_nombre(col_maestro)
        if norm_maestro in normalizados_fuente:
            mapeo[col_maestro] = normalizados_fuente[norm_maestro]
        else:
            # Variantes comunes (puedes ampliar esta lista)
            variantes = {
                "fecha de compra": ["fecha compra", "fecha_compra", "fechacompra"],
                "n°": ["n", "numero", "num", "nº", "no"],
                "n° op": ["n op", "n op.", "numero op", "nop", "n° operacion"],
                "id blotter": ["id blotter", "id_blotter", "blotter"],
                "apellido paterno": ["apellido paterno", "ap paterno", "paterno"],
                "apellido materno": ["apellido materno", "ap materno", "materno"],
                "nombres": ["nombres", "nombre completo"],
                "rut": ["rut", "run", "rut cliente"],
                "dv": ["dv", "digito verificador"],
                "fecha de emision": ["fecha emision", "fecha_emision", "fecha emisión", "fecha de emision", "fecha de suscripcion"],
                "monto credito + cap (uf)": [
                    "monto credito + cap (uf)", "monto credito - cap (uf)",
                    "monto credito uf", "credito uf", "monto cap",
                    "monto credito+cap (uf)", "monto credito + cap(uf)",
                ],
                "subsidio": ["subsidio"],
                "pie": ["pie", "pie inicial"],
                "valor vivienda": ["valor vivienda", "valor_vivienda", "valor propiedad"],
                "morosidad": ["morosidad", "dias morosidad"],
                "n° cuotas": ["n cuotas", "numero cuotas", "cuotas", "nº cuotas"],
                "cuota mes": ["cuota mes", "cuota", "cuota mensual", "primera cuota a endosar"],
                "tasa arriendo o compra": ["tasa arriendo", "tasa compra", "tasa de compra", "tasa anual de emision", "tasa"],
                "fecha 1er aporte": ["fecha primer aporte", "1er aporte", "fecha 1er aporte", "fecha 1er aporte a endosar"],
                "fecha last aporte": ["fecha ultimo aporte", "fecha ultimo aporte a endosar", "last aporte", "fecha last aporte"],
                "fecha corte": ["fecha corte", "fecha de corte", "corte"],
                "saldo insoluto teorico al 31-07-2019": ["saldo insoluto teorico", "saldo insoluto", "saldo teorico", "saldo 31-07-2019"],
                "tasacion": ["tasacion", "tasación", "valor tasacion", "tasacion de la propiedad"],
                "precio venta/tasacion": ["precio venta", "precio tasacion", "precio venta tasacion", "precio venta/tasacion"],
                "tasa venta": ["tasa venta", "tasa_venta", "tasa de venta", "tasa de endoso", "tasa anual de endoso", "tasa endoso"],
                "dif. tasa": ["dif tasa", "diferencia tasa", "dif tasa"],
                "monto dividendo": ["monto dividendo", "dividendo", "monto dividendo"],
                "div/renta": ["div/renta", "divrenta", "dividendo/renta", "dividendo/ renta"],
                "divrenta": ["div/renta", "divrenta", "dividendo/renta", "dividendo/ renta"],
                "carga financiera": ["carga financiera", "carga_financiera", "carga financiera/ renta"],
                "direccion": ["direccion", "dirección", "domicilio", "direccion de la propiedad"],
                "comuna": ["comuna"],
            }
            clave = norm_maestro
            if clave in variantes:
                for variante in variantes[clave]:
                    if variante in normalizados_fuente:
                        mapeo[col_maestro] = normalizados_fuente[variante]
                        break
            # Fallback: columna del fuente cuyo nombre normalizado empieza por el del maestro
            # (ej. "Fecha 1er Aporte" -> "Fecha 1er Aporte a endosar"; "1.er" ya unificado en normalizar)
            if col_maestro not in mapeo:
                candidatos = [(k, normalizados_fuente[k]) for k in normalizados_fuente if k.startswith(norm_maestro)]
                # Prefijos alternativos (ej. fuente tiene "Último" y maestro "Last")
                prefijos_alternativos = {
                    "fecha last aporte": ["fecha ultimo aporte"],
                    "fecha corte": ["fecha de corte", "corte"],
                }
                if norm_maestro in prefijos_alternativos:
                    for prefijo in prefijos_alternativos[norm_maestro]:
                        candidatos.extend(
                            [(k, normalizados_fuente[k]) for k in normalizados_fuente if k.startswith(prefijo)]
                        )
                if candidatos:
                    # Elegir la coincidencia más específica (nombre más largo)
                    mejor = max(candidatos, key=lambda x: len(x[0]))
                    mapeo[col_maestro] = mejor[1]
            # Fallback columnas de fecha (1er/Last Aporte, Corte): buscar por contenido del nombre
            if col_maestro not in mapeo:
                if norm_maestro == "fecha 1er aporte":
                    candidatos = [(k, normalizados_fuente[k]) for k in normalizados_fuente if "fecha" in k and "aporte" in k and ("1er" in k or "1.er" in k or "primer" in k)]
                elif norm_maestro == "fecha last aporte":
                    candidatos = [(k, normalizados_fuente[k]) for k in normalizados_fuente if "fecha" in k and "aporte" in k and ("last" in k or "ultimo" in k)]
                elif norm_maestro == "fecha corte":
                    candidatos = [(k, normalizados_fuente[k]) for k in normalizados_fuente if ("fecha" in k and "corte" in k) or k == "corte" or k.startswith("corte ")]
                else:
                    candidatos = []
                if candidatos:
                    mejor = max(candidatos, key=lambda x: len(x[0]))
                    mapeo[col_maestro] = mejor[1]
            # Fallback Tasacion: columna en fuente puede llamarse "Tasación", "Valor Tasación", etc.
            if col_maestro not in mapeo and norm_maestro == "tasacion":
                candidatos = [
                    (k, normalizados_fuente[k]) for k in normalizados_fuente
                    if "tasacion" in k and "precio" not in k
                ]
                if candidatos:
                    mejor = max(candidatos, key=lambda x: len(x[0]))
                    mapeo[col_maestro] = mejor[1]
            # Fallback Monto Crédito + Cap (UF): nombre en fuente puede variar (+/- , espacios)
            if col_maestro not in mapeo and norm_maestro == "monto credito + cap (uf)":
                candidatos = [
                    (k, normalizados_fuente[k]) for k in normalizados_fuente
                    if "monto credito" in k and "cap" in k and "uf" in k
                ]
                if candidatos:
                    mejor = max(candidatos, key=lambda x: len(x[0]))
                    mapeo[col_maestro] = mejor[1]
            # Fallback Saldo insoluto: la cabecera en el fuente puede ser solo la fecha "31-07-2019"
            if col_maestro not in mapeo and "saldo insoluto" in norm_maestro:
                candidatos = [
                    (k, normalizados_fuente[k]) for k in normalizados_fuente
                    if ("saldo" in k and "insoluto" in k)
                    or "31-07-2019" in k or "31/07/2019" in k
                ]
                if candidatos:
                    mejor = max(candidatos, key=lambda x: len(x[0]))
                    mapeo[col_maestro] = mejor[1]
            # Fallback por palabra clave: buscar columnas cuyo nombre CONTIENE las
            # palabras esenciales (para nombres variables como "Morosidad al 10-03-21")
            if col_maestro not in mapeo:
                palabras_clave = {
                    "morosidad": [["morosidad"]],
                    "saldo insoluto teorico al 31-07-2019": [["saldo", "insoluto"]],
                    "fecha de emision": [["fecha", "emision"], ["fecha", "suscripcion"]],
                    "tasa arriendo o compra": [["tasa", "compra"], ["tasa", "arriendo"], ["tasa", "emision"]],
                    "tasa venta": [["tasa", "venta"], ["tasa", "endoso"]],
                    "precio venta/tasacion": [["precio", "venta"]],
                    "carga financiera": [["carga", "financiera"]],
                    "monto credito + cap (uf)": [["monto", "credito"]],
                }
                if norm_maestro in palabras_clave:
                    for grupo_kw in palabras_clave[norm_maestro]:
                        candidatos = [
                            (k, normalizados_fuente[k]) for k in normalizados_fuente
                            if all(kw in k for kw in grupo_kw)
                        ]
                        if candidatos:
                            mejor = max(candidatos, key=lambda x: len(x[0]))
                            mapeo[col_maestro] = mejor[1]
                            break
            # Fallback por índice: columnas del fuente con encabezado vacío (Unnamed)
            if col_maestro not in mapeo and col_maestro in COLUMNAS_POR_INDICE_FUENTE:
                idx = COLUMNAS_POR_INDICE_FUENTE[col_maestro]
                if 0 <= idx < len(df_fuente.columns):
                    mapeo[col_maestro] = df_fuente.columns[idx]
    return mapeo


def imprimir_reporte_mapeo(mapeo: dict, columnas_fuente: list, columnas_fuente_adicionales_usadas: tuple = ()) -> None:
    """
    Imprime qué columnas del archivo fuente se mapearon a cada columna del maestro
    y cuáles del maestro quedaron sin mapear. También lista columnas del fuente no usadas.
    columnas_fuente_adicionales_usadas: columnas usadas por lógica (ej. columnas con fecha para VPN).
    """
    columnas_fuente = [c for c in columnas_fuente if isinstance(c, str) or not pd.isna(c)]
    usadas = set(mapeo.values()) | set(c for c in columnas_fuente_adicionales_usadas if c is not None)

    print("\n" + "=" * 60)
    print("REPORTE DE MAPEO DE COLUMNAS")
    print("=" * 60)
    print("\n--- Columnas del maestro (mapeadas) ---")
    for col_maestro in COLUMNAS_MAESTRO:
        if col_maestro in mapeo:
            col_fuente = mapeo[col_maestro]
            print(f"  [OK] {col_maestro!r}")
            print(f"       <- {col_fuente!r}")
        elif col_maestro in COLUMNAS_CALCULADAS:
            print(f"  [CALCULADO] {col_maestro!r} (desde {COLUMNAS_CALCULADAS[col_maestro]!r})")
        elif col_maestro in COLUMNAS_DESDE_FECHA:
            print(f"  [DESDE COL. FECHA] {col_maestro!r} (1ª/2ª col. del fuente con cabecera fecha)")
        elif col_maestro in COLUMNAS_DESDE_BASE:
            print(f"  [DESDE HOJA BASE] {col_maestro!r} <- {COLUMNAS_DESDE_BASE[col_maestro]!r}")
        else:
            print(f"  [SIN MAPEAR] {col_maestro!r} (quedará vacío en filas nuevas)")
    print("\n--- Columnas del archivo fuente no usadas ---")
    print("  (índice = posición 0-based para COLUMNAS_POR_INDICE_FUENTE en cargar_a_maestro.py)")
    sin_uso_con_indice = [(i, c) for i, c in enumerate(columnas_fuente) if c not in usadas]
    if sin_uso_con_indice:
        for idx, c in sin_uso_con_indice:
            print(f"  - índice {idx}: {c!r}")
    else:
        print("  (ninguna; todas las columnas del fuente se usaron)")
    print("=" * 60 + "\n")


def aplicar_columnas_calculadas(df_out: pd.DataFrame) -> None:
    """
    Rellena columnas del maestro que no vienen del archivo fuente,
    calculándolas a partir de otras columnas. Modifica df_out in-place.
    """
    # ID Blotter = todos los dígitos de N° OP antes de la primera letra
    if "N° OP" in df_out.columns:
        id_blotter = df_out["N° OP"].map(extraer_id_blotter_desde_n_op)
        df_out["ID Blotter"] = id_blotter.replace("", pd.NA)

    # Dif. Tasa = Tasa Arriendo o Compra - Tasa Venta
    if "Tasa Arriendo o Compra" in df_out.columns and "Tasa Venta" in df_out.columns:
        tasa_arriendo = pd.to_numeric(df_out["Tasa Arriendo o Compra"], errors="coerce")
        tasa_venta = pd.to_numeric(df_out["Tasa Venta"], errors="coerce")
        df_out["Dif. Tasa"] = tasa_arriendo - tasa_venta


def rellenar_vpn_desde_columnas_fecha(
    df_out: pd.DataFrame,
    df_fuente: pd.DataFrame,
    col_1ra: str = None,
    col_2da: str = None,
) -> None:
    """
    Rellena 'VPN 1ra fecha', 'VPN 2da fecha' y 'Precio -1UF' desde el fuente:
    - VPN 1ra/2da = columnas cuyo encabezado es la 1ª y 2ª fecha (orden cronológico).
    - Precio -1UF = valor de la columna de la 2ª fecha menos 1.
    """
    if col_1ra is None and col_2da is None:
        col_1ra, col_2da = obtener_columnas_fecha_vpn(df_fuente)
    if col_1ra is not None and "VPN 1ra fecha" in df_out.columns:
        df_out["VPN 1ra fecha"] = df_fuente[col_1ra].values
    if col_2da is not None and "VPN 2da fecha" in df_out.columns:
        df_out["VPN 2da fecha"] = df_fuente[col_2da].values
    if col_2da is not None and "Precio -1UF" in df_out.columns:
        serie = df_fuente[col_2da].astype(str).str.replace(",", ".", regex=False)
        valores = pd.to_numeric(serie, errors="coerce")
        df_out["Precio -1UF"] = valores - 1


def _convertir_columna_a_fecha(serie: pd.Series) -> pd.Series:
    """Convierte una serie a datetime; acepta texto dd/mm/yyyy, números serial Excel y datetime."""
    if serie.empty:
        return serie
    # Si ya es datetime64, solo normalizar
    if pd.api.types.is_datetime64_any_dtype(serie):
        return pd.to_datetime(serie, errors="coerce").dt.normalize()
    # Números que parecen serial Excel (días desde 1899-12-30): típicamente 1000–100000
    numeric = pd.to_numeric(serie, errors="coerce")
    if numeric.notna().any():
        sample = numeric.dropna()
        if (sample >= 1000).all() and (sample <= 100000).all():
            try:
                fechas = pd.TimedeltaIndex(numeric.astype(float), unit="d") + pd.Timestamp("1899-12-30")
                return pd.Series(fechas.normalize(), index=serie.index)
            except (ValueError, TypeError):
                pass
    # Texto o resto: convertir (dayfirst para formato chileno dd/mm/yyyy)
    return pd.to_datetime(serie, dayfirst=True, errors="coerce").dt.normalize()


def dataframe_fuente_a_formato_maestro(
    df_fuente: pd.DataFrame,
    mapeo: dict = None,
    col_vpn_1ra: str = None,
    col_vpn_2da: str = None,
) -> pd.DataFrame:
    """Construye un DataFrame con las columnas del maestro, rellenando desde el fuente según mapeo."""
    if mapeo is None:
        mapeo = mapear_columnas_fuente_a_maestro(df_fuente)
    df_out = pd.DataFrame(columns=COLUMNAS_MAESTRO)
    for col_maestro in COLUMNAS_MAESTRO:
        if col_maestro in mapeo:
            df_out[col_maestro] = df_fuente[mapeo[col_maestro]].values
        else:
            df_out[col_maestro] = pd.NA
    for col in COLUMNAS_FECHA:
        if col in df_out.columns:
            df_out[col] = _convertir_columna_a_fecha(df_out[col])
    rellenar_vpn_desde_columnas_fecha(df_out, df_fuente, col_1ra=col_vpn_1ra, col_2da=col_vpn_2da)
    return df_out


def _leer_valo_con_filas_2_y_3(ruta_fuente: str) -> pd.DataFrame:
    """
    Lee la hoja Valo sin usar una fila fija como encabezado; construye los nombres
    de columna desde las filas 2 y 3 (Excel): para cada columna usa el valor de
    la fila 3 si no está vacío, si no el de la fila 2. Los datos empiezan en la fila 4.
    """
    try:
        df_raw = pd.read_excel(ruta_fuente, sheet_name=HOJA_VALO, header=None)
    except ValueError:
        df_raw = pd.read_excel(ruta_fuente, sheet_name=0, header=None)
    if df_raw.empty or len(df_raw) < 3:
        return pd.DataFrame()
    # Excel fila 2 = índice 1, Excel fila 3 = índice 2, datos desde fila 4 = índice 3
    row2 = df_raw.iloc[1]
    row3 = df_raw.iloc[2]
    headers = []
    for i in range(len(df_raw.columns)):
        v3 = row3.iloc[i]
        v2 = row2.iloc[i]
        if pd.notna(v3) and str(v3).strip():
            headers.append(str(v3).strip())
        elif pd.notna(v2) and str(v2).strip():
            headers.append(str(v2).strip())
        else:
            headers.append(f"Unnamed: {i}")
    df_fuente = df_raw.iloc[3:].copy()
    df_fuente.columns = headers
    df_fuente.reset_index(drop=True, inplace=True)
    return df_fuente


def _leer_maestro_como_dataframe(ruta_maestro: str) -> pd.DataFrame:
    """
    Lee el maestro construyendo los encabezados desde filas 1 y 2 (pueden estar
    en celdas combinadas). Para cada columna usa el valor de fila 2 si existe,
    si no el de fila 1. Los datos empiezan en fila 3.
    Devuelve un DataFrame con columnas estandarizadas a COLUMNAS_MAESTRO.
    """
    df_raw = pd.read_excel(ruta_maestro, sheet_name=0, header=None)
    if df_raw.empty or len(df_raw) < 3:
        return pd.DataFrame(columns=COLUMNAS_MAESTRO)
    row1 = df_raw.iloc[0]
    row2 = df_raw.iloc[1]
    n_cols = min(len(df_raw.columns), len(COLUMNAS_MAESTRO) + 10)
    headers = []
    for i in range(n_cols):
        v2 = row2.iloc[i] if i < len(row2) else None
        v1 = row1.iloc[i] if i < len(row1) else None
        if pd.notna(v2) and str(v2).strip():
            headers.append(str(v2).strip())
        elif pd.notna(v1) and str(v1).strip():
            headers.append(str(v1).strip())
        else:
            headers.append(f"Unnamed: {i}")
    df = df_raw.iloc[2:, :n_cols].copy()
    df.columns = headers
    df.reset_index(drop=True, inplace=True)
    return _maestro_a_columnas_estandar(df)


def _es_na(val) -> bool:
    """Verifica si un valor es NA/NaN/NaT de forma segura."""
    try:
        return pd.isna(val)
    except (ValueError, TypeError):
        return False


_ALINEACION_CENTRO = Alignment(horizontal="center", vertical="center")


def _escribir_celda(ws, fila, col_idx, col_name, val, columnas_porcentaje, columnas_fecha):
    """Escribe un valor en una celda de Excel con el formato correcto y centrado."""
    cell = ws.cell(row=fila, column=col_idx)

    if _es_na(val):
        cell.alignment = _ALINEACION_CENTRO
        return

    if col_name in columnas_porcentaje:
        num_val = val if isinstance(val, (int, float)) else pd.to_numeric(val, errors="coerce")
        if isinstance(num_val, (int, float)) and not _es_na(num_val):
            if abs(num_val) > 1:
                num_val = num_val / 100
            cell.value = num_val
            cell.number_format = "0.00%"
        cell.alignment = _ALINEACION_CENTRO
        return

    if col_name in columnas_fecha:
        if isinstance(val, pd.Timestamp):
            cell.value = val.to_pydatetime()
        elif hasattr(val, "to_pydatetime"):
            cell.value = val.to_pydatetime()
        else:
            try:
                cell.value = pd.to_datetime(val, dayfirst=True).to_pydatetime()
            except Exception:
                cell.value = val
        cell.number_format = "dd/mm/yyyy"
        cell.alignment = _ALINEACION_CENTRO
        return

    cell.value = val
    cell.alignment = _ALINEACION_CENTRO


def cargar_y_agregar_a_maestro(ruta_fuente: str, ruta_maestro: str, fecha_compra: str = None) -> None:
    """
    Lee el archivo fuente, lo mapea al formato maestro y AGREGA las filas nuevas
    al maestro existente sin tocar las filas ya presentes.
    Evita duplicados por Rut + Fecha de emisión.
    Si se indica fecha_compra, se usa para todas las filas nuevas.
    """
    if not os.path.isfile(ruta_fuente):
        raise FileNotFoundError(f"No se encontró el archivo fuente: {ruta_fuente}")

    # Leer hoja Valo (datos principales)
    if USAR_FILAS_2_Y_3_VALO:
        df_fuente = _leer_valo_con_filas_2_y_3(ruta_fuente)
    else:
        try:
            df_fuente = pd.read_excel(ruta_fuente, sheet_name=HOJA_VALO, header=FILA_ENCABEZADO_VALO)
        except ValueError:
            df_fuente = pd.read_excel(ruta_fuente, sheet_name=0, header=FILA_ENCABEZADO_VALO)
    if df_fuente.empty:
        print("El archivo fuente no tiene filas en la hoja de datos. No se agrega nada.")
        return

    # Leer hoja Base (para columnas adicionales); nombres de columna en fila 1
    df_base = None
    try:
        df_base = pd.read_excel(ruta_fuente, sheet_name=HOJA_BASE, header=FILA_ENCABEZADO_BASE)
    except ValueError:
        pass

    # Mapeo y reporte
    mapeo = mapear_columnas_fuente_a_maestro(df_fuente)
    columnas_ya_mapeadas = set(mapeo.values())
    col_1ra, col_2da = obtener_columnas_fecha_vpn(df_fuente, columnas_excluir=columnas_ya_mapeadas)
    imprimir_reporte_mapeo(mapeo, list(df_fuente.columns), columnas_fuente_adicionales_usadas=(col_1ra, col_2da))

    # Convertir al formato del maestro (desde Valo)
    df_nuevo = dataframe_fuente_a_formato_maestro(
        df_fuente, mapeo=mapeo, col_vpn_1ra=col_1ra, col_vpn_2da=col_2da
    )

    # Rellenar 3 columnas desde la hoja Base (enlace por Rut u otra llave)
    if df_base is not None and not df_base.empty and COLUMNAS_DESDE_BASE:
        rellenar_desde_hoja_base(df_nuevo, df_base)

    # Calcular columnas derivadas DESPUÉS de rellenar desde Base
    aplicar_columnas_calculadas(df_nuevo)

    # Saldo insoluto: si está vacío, rellenar con 0
    col_saldo = "Saldo insoluto Teorico al 31-07-2019"
    if col_saldo in df_nuevo.columns:
        df_nuevo[col_saldo] = pd.to_numeric(df_nuevo[col_saldo], errors="coerce").fillna(0)

    # Fecha de compra como input
    if fecha_compra is not None and fecha_compra.strip():
        try:
            df_nuevo["Fecha de compra"] = pd.to_datetime(fecha_compra.strip(), dayfirst=True)
        except (ValueError, TypeError):
            df_nuevo["Fecha de compra"] = fecha_compra.strip()

    # Normalizar todas las columnas de fecha del nuevo DataFrame
    for col in COLUMNAS_FECHA:
        if col in df_nuevo.columns:
            df_nuevo[col] = pd.to_datetime(df_nuevo[col], dayfirst=True, errors="coerce")

    COLUMNAS_PORCENTAJE = ("Tasa Arriendo o Compra", "Tasa Venta", "Dif. Tasa", "Precio Venta/Tasación")

    maestro_existe = os.path.isfile(ruta_maestro)

    if maestro_existe:
        # --- MODO APPEND: solo agregar filas nuevas, nunca tocar las existentes ---

        # Leer maestro SOLO para obtener llaves de deduplicación.
        # Las cabeceras pueden estar en fila 1 o 2 (celdas combinadas).
        df_maestro_std = _leer_maestro_como_dataframe(ruta_maestro)
        n_antes = len(df_maestro_std)

        llaves_existentes = set()
        for _, r in df_maestro_std.iterrows():
            rut = _rut_valor_a_str(r["Rut"]) if not _es_na(r.get("Rut")) else ""
            fec = _normalizar_fecha_str(r.get("Fecha de emisión"))
            if rut and fec:
                llaves_existentes.add((rut, fec))

        def _es_duplicado(r):
            rut = _rut_valor_a_str(r["Rut"]) if not _es_na(r.get("Rut")) else ""
            fec = _normalizar_fecha_str(r.get("Fecha de emisión"))
            if not rut or not fec:
                return False
            return (rut, fec) in llaves_existentes

        mask_dup = df_nuevo.apply(_es_duplicado, axis=1)
        df_a_agregar = df_nuevo[~mask_dup].copy()
        n_duplicados = int(mask_dup.sum())

        if df_a_agregar.empty:
            print("No hay filas nuevas para agregar (todas ya existen en el maestro).")
            return

        # Abrir el archivo existente y agregar al final
        wb = load_workbook(ruta_maestro)
        ws = wb.active

        # Leer cabeceras: pueden estar en fila 1, fila 2 o celdas combinadas
        n_maestro_cols = len(COLUMNAS_MAESTRO)
        max_scan = min(ws.max_column, n_maestro_cols + 10)
        col_pos = {}
        for ci in range(1, max_scan + 1):
            v1 = ws.cell(row=1, column=ci).value
            v2 = ws.cell(row=2, column=ci).value
            val = v2 if v2 is not None else v1
            if val is not None:
                norm_val = normalizar_nombre(str(val))
                for cm in COLUMNAS_MAESTRO:
                    if cm not in col_pos and normalizar_nombre(cm) == norm_val:
                        col_pos[cm] = ci
                        break
        # Fallback posicional: si una columna no se mapeó por nombre, usar su
        # posición en COLUMNAS_MAESTRO (el maestro tiene ese orden exacto)
        for i, cm in enumerate(COLUMNAS_MAESTRO):
            if cm not in col_pos and (i + 1) <= max_scan:
                col_pos[cm] = i + 1

        print(f"  Columnas mapeadas en maestro: {len(col_pos)}/{n_maestro_cols}")

        siguiente_fila = ws.max_row + 1

        # Obtener el último N° del maestro para continuar la enumeración
        ultimo_n = 0
        if "N°" in col_pos:
            col_n = col_pos["N°"]
            for fila_r in range(ws.max_row, 2, -1):
                val_n = ws.cell(row=fila_r, column=col_n).value
                if val_n is not None:
                    try:
                        ultimo_n = int(val_n)
                    except (ValueError, TypeError):
                        pass
                    break

        for i, (_, row) in enumerate(df_a_agregar.iterrows()):
            fila_excel = siguiente_fila + i
            for col_name in COLUMNAS_MAESTRO:
                if col_name not in col_pos:
                    continue
                if col_name == "N°":
                    val = ultimo_n + i + 1
                else:
                    val = row.get(col_name)
                _escribir_celda(
                    ws, fila_excel, col_pos[col_name], col_name, val,
                    COLUMNAS_PORCENTAJE, COLUMNAS_FECHA,
                )

        wb.save(ruta_maestro)

        print(f"Maestro actualizado: {ruta_maestro}")
        print(f"  Filas en maestro antes: {n_antes}")
        print(f"  Filas nuevas leídas:    {len(df_nuevo)}")
        print(f"  Duplicados (ya existían): {n_duplicados}")
        print(f"  Filas agregadas:        {len(df_a_agregar)}")
        print(f"  Total filas en maestro: {n_antes + len(df_a_agregar)}")

    else:
        # --- MODO CREAR: no existe maestro, crear uno nuevo ---
        # Asignar N° correlativo empezando desde 1
        if "N°" in df_nuevo.columns:
            df_nuevo["N°"] = range(1, len(df_nuevo) + 1)

        with pd.ExcelWriter(ruta_maestro, engine="openpyxl") as writer:
            df_nuevo.to_excel(writer, index=False, sheet_name="Sheet1", startrow=1)
            ws = writer.sheets["Sheet1"]
            n_filas = len(df_nuevo)
            fila_inicio = 3
            fila_fin = fila_inicio + n_filas
            n_cols = len(df_nuevo.columns)
            for col_name in COLUMNAS_PORCENTAJE:
                if col_name not in df_nuevo.columns:
                    continue
                col_idx = df_nuevo.columns.get_loc(col_name) + 1
                col_letter = get_column_letter(col_idx)
                for row in range(fila_inicio, fila_fin):
                    cell = ws[f"{col_letter}{row}"]
                    val = cell.value
                    if isinstance(val, (int, float)) and not _es_na(val):
                        if abs(val) > 1:
                            cell.value = val / 100
                        cell.number_format = "0.00%"
            for col_name in COLUMNAS_FECHA:
                if col_name not in df_nuevo.columns:
                    continue
                col_idx = df_nuevo.columns.get_loc(col_name) + 1
                col_letter = get_column_letter(col_idx)
                for row in range(fila_inicio, fila_fin):
                    ws[f"{col_letter}{row}"].number_format = "dd/mm/yyyy"
            # Centrar todas las celdas de datos
            for row in range(fila_inicio, fila_fin):
                for ci in range(1, n_cols + 1):
                    ws.cell(row=row, column=ci).alignment = _ALINEACION_CENTRO

        print(f"Maestro creado: {ruta_maestro}")
        print(f"  Filas: {len(df_nuevo)}")


def main():
    if len(sys.argv) < 2:
        print("Uso: python cargar_a_maestro.py <ruta_archivo_fuente.xlsx> [ruta_maestro.xlsx] [fecha_compra]")
        print("  Si no se indica ruta_maestro, se usa 'maestro.xlsx' en la misma carpeta del script.")
        print("  Si no se indica fecha_compra, se pedirá por consola (ej. 2024-01-15).")
        sys.exit(1)

    ruta_fuente = sys.argv[1].strip()
    if len(sys.argv) >= 3:
        ruta_maestro = sys.argv[2].strip()
    else:
        carpeta = os.path.dirname(os.path.abspath(__file__))
        ruta_maestro = os.path.join(carpeta, "maestro.xlsx")

    # Fecha de compra: por argumento o por input
    if len(sys.argv) >= 4:
        fecha_compra = sys.argv[3].strip()
    else:
        fecha_compra = input("Fecha de compra (ej. 2024-01-15 o 15/01/2024): ").strip()

    try:
        cargar_y_agregar_a_maestro(ruta_fuente, ruta_maestro, fecha_compra=fecha_compra or None)
    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
